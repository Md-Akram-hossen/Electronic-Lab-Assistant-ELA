from __future__ import annotations

import datetime as dt
import logging
import sqlite3
from io import BytesIO
from pathlib import Path
from threading import Thread
from typing import Optional

from zoneinfo import ZoneInfo

from .config import get_settings
from .google_client import send_simple_email
from .parsing import normalize_component_name

try:
    import openpyxl
    from openpyxl import Workbook
except Exception:
    openpyxl = None
    Workbook = None

try:
    from dateutil.relativedelta import relativedelta
except Exception:
    relativedelta = None


def connect_lab() -> sqlite3.Connection:
    settings = get_settings()
    conn = sqlite3.connect(settings.lab_db)
    conn.execute("PRAGMA journal_mode=WAL;")
    conn.execute("PRAGMA synchronous=NORMAL;")
    return conn


def _column_exists(conn: sqlite3.Connection, table: str, column: str) -> bool:
    rows = conn.execute(f"PRAGMA table_info({table})").fetchall()
    return any((row[1] or "").lower() == column.lower() for row in rows)


def ensure_lab_db() -> None:
    conn = connect_lab()
    cur = conn.cursor()
    cur.execute(
        '''
        CREATE TABLE IF NOT EXISTS components (
            name TEXT PRIMARY KEY COLLATE NOCASE,
            quantity INTEGER NOT NULL,
            location TEXT NOT NULL,
            locker INTEGER NOT NULL
        )
        '''
    )
    cur.execute(
        '''
        CREATE TABLE IF NOT EXISTS students (
            student_id TEXT PRIMARY KEY,
            name TEXT NOT NULL,
            dob TEXT,
            email TEXT
        )
        '''
    )
    cur.execute(
        '''
        CREATE TABLE IF NOT EXISTS borrow_log (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            ts DATETIME DEFAULT CURRENT_TIMESTAMP,
            student_id TEXT NOT NULL,
            student_name TEXT NOT NULL,
            item TEXT NOT NULL,
            qty INTEGER NOT NULL,
            remaining INTEGER NOT NULL
        )
        '''
    )
    cur.execute(
        '''
        CREATE TABLE IF NOT EXISTS return_log (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            ts DATETIME DEFAULT CURRENT_TIMESTAMP,
            student_id TEXT NOT NULL,
            student_name TEXT NOT NULL,
            item TEXT NOT NULL,
            qty INTEGER NOT NULL,
            new_total INTEGER NOT NULL
        )
        '''
    )
    cur.execute(
        '''
        CREATE TABLE IF NOT EXISTS borrow_txn (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            borrow_ts TEXT NOT NULL,
            due_ts TEXT NOT NULL,
            student_id TEXT NOT NULL,
            student_name TEXT NOT NULL,
            student_email TEXT,
            item TEXT NOT NULL,
            qty INTEGER NOT NULL,
            returned_qty INTEGER NOT NULL DEFAULT 0,
            last_reminder_ts TEXT,
            reminder_count INTEGER NOT NULL DEFAULT 0
        )
        '''
    )
    cur.execute("CREATE INDEX IF NOT EXISTS idx_components_locker ON components(locker)")
    cur.execute("CREATE INDEX IF NOT EXISTS idx_students_name ON students(name)")
    cur.execute("CREATE INDEX IF NOT EXISTS idx_borrow_txn_due ON borrow_txn(due_ts)")
    cur.execute("CREATE INDEX IF NOT EXISTS idx_borrow_txn_student_item ON borrow_txn(student_id, item)")
    conn.commit()
    conn.close()


def seed_defaults_if_empty() -> None:
    conn = connect_lab()
    cur = conn.cursor()
    count = cur.execute("SELECT COUNT(*) FROM components").fetchone()[0]
    if not count:
        cur.executemany(
            "INSERT INTO components(name, quantity, location, locker) VALUES(?, ?, ?, ?)",
            [
                ("arduino", 10, "Room L30 Locker 1", 1),
                ("servo", 20, "Room L30 Locker 2", 2),
                ("seven segment", 10, "Room L30 Locker 3", 3),
            ],
        )
    count = cur.execute("SELECT COUNT(*) FROM students").fetchone()[0]
    if not count:
        cur.executemany(
            "INSERT INTO students(student_id, name, dob, email) VALUES(?, ?, ?, ?)",
            [
                ("2234", "Akram", "01-01-2000", "akram@example.com"),
                ("1122", "Alice", "11-12-2001", "alice@example.com"),
                ("3344", "Bob", "13-02-2004", "bob@example.com"),
            ],
        )
    conn.commit()
    conn.close()


def get_component(name: str):
    conn = connect_lab()
    row = conn.execute(
        "SELECT name, quantity, location, locker FROM components WHERE name=? COLLATE NOCASE",
        (normalize_component_name(name),),
    ).fetchone()
    conn.close()
    return row


def list_all_components():
    conn = connect_lab()
    rows = conn.execute("SELECT name, quantity, location, locker FROM components ORDER BY name").fetchall()
    conn.close()
    return rows


def resolve_component_name(text: str) -> str:
    candidate = normalize_component_name(text)
    if get_component(candidate):
        return candidate
    try:
        import difflib
        names = [row[0] for row in list_all_components()]
        best = difflib.get_close_matches(candidate, names, n=1, cutoff=0.84)
        return best[0] if best else candidate
    except Exception:
        return candidate


def get_student(student_id: str):
    conn = connect_lab()
    row = conn.execute(
        "SELECT student_id, name, email FROM students WHERE student_id=?",
        (student_id,),
    ).fetchone()
    conn.close()
    return row


def atomic_borrow(item_name: str, borrow_qty: int):
    item_name = normalize_component_name(item_name)
    conn = connect_lab()
    conn.isolation_level = None
    cur = conn.cursor()
    try:
        cur.execute("BEGIN IMMEDIATE")
        row = cur.execute(
            "SELECT quantity, location, locker FROM components WHERE name=? COLLATE NOCASE",
            (item_name,),
        ).fetchone()
        if not row:
            cur.execute("ROLLBACK")
            return False, None, "not-found"
        current, location, locker = int(row[0]), row[1], int(row[2])
        if borrow_qty <= 0:
            cur.execute("ROLLBACK")
            return False, None, "invalid-quantity"
        if current < borrow_qty:
            cur.execute("ROLLBACK")
            return False, None, "insufficient-stock"
        new_qty = current - borrow_qty
        cur.execute("UPDATE components SET quantity=? WHERE name=? COLLATE NOCASE", (new_qty, item_name))
        cur.execute("COMMIT")
        return True, new_qty, {"location": location, "locker": locker}
    except Exception:
        try:
            cur.execute("ROLLBACK")
        except Exception:
            pass
        logging.exception("atomic_borrow failed")
        return False, None, "error"
    finally:
        conn.close()


def atomic_return(item_name: str, return_qty: int):
    item_name = normalize_component_name(item_name)
    conn = connect_lab()
    conn.isolation_level = None
    cur = conn.cursor()
    try:
        cur.execute("BEGIN IMMEDIATE")
        row = cur.execute(
            "SELECT quantity, location, locker FROM components WHERE name=? COLLATE NOCASE",
            (item_name,),
        ).fetchone()
        if not row:
            cur.execute("ROLLBACK")
            return False, None, "not-found"
        current, location, locker = int(row[0]), row[1], int(row[2])
        if return_qty <= 0:
            cur.execute("ROLLBACK")
            return False, None, "invalid-quantity"
        new_total = current + return_qty
        cur.execute("UPDATE components SET quantity=? WHERE name=? COLLATE NOCASE", (new_total, item_name))
        cur.execute("COMMIT")
        return True, new_total, {"location": location, "locker": locker}
    except Exception:
        try:
            cur.execute("ROLLBACK")
        except Exception:
            pass
        logging.exception("atomic_return failed")
        return False, None, "error"
    finally:
        conn.close()


def log_borrow(student_id: str, student_name: str, item: str, qty: int, remaining: int) -> None:
    conn = connect_lab()
    conn.execute(
        "INSERT INTO borrow_log(student_id, student_name, item, qty, remaining) VALUES(?,?,?,?,?)",
        (student_id, student_name, normalize_component_name(item), qty, remaining),
    )
    conn.commit()
    conn.close()


def log_return(student_id: str, student_name: str, item: str, qty: int, new_total: int) -> None:
    conn = connect_lab()
    conn.execute(
        "INSERT INTO return_log(student_id, student_name, item, qty, new_total) VALUES(?,?,?,?,?)",
        (student_id, student_name, normalize_component_name(item), qty, new_total),
    )
    conn.commit()
    conn.close()


def _utcnow() -> dt.datetime:
    return dt.datetime.now(dt.timezone.utc)


def _add_months(start: dt.datetime, months: int) -> dt.datetime:
    if relativedelta is None:
        return start + dt.timedelta(days=30 * months)
    return start + relativedelta(months=+months)


def add_borrow_txn(student_id: str, student_name: str, student_email: str, item: str, qty: int):
    settings = get_settings()
    borrow_dt = _utcnow()
    due_dt = _add_months(borrow_dt, settings.borrow_due_months)
    conn = connect_lab()
    conn.execute(
        '''
        INSERT INTO borrow_txn(
            borrow_ts, due_ts, student_id, student_name, student_email, item, qty, returned_qty
        ) VALUES(?,?,?,?,?,?,?,0)
        ''',
        (
            borrow_dt.isoformat(timespec="seconds"),
            due_dt.isoformat(timespec="seconds"),
            student_id,
            student_name,
            student_email,
            normalize_component_name(item),
            qty,
        ),
    )
    conn.commit()
    conn.close()
    return borrow_dt, due_dt


def apply_return_fifo(student_id: str, item: str, return_qty: int):
    conn = connect_lab()
    conn.isolation_level = None
    cur = conn.cursor()
    applied = 0
    remaining = int(return_qty)
    try:
        cur.execute("BEGIN IMMEDIATE")
        rows = cur.execute(
            '''
            SELECT id, qty, returned_qty
            FROM borrow_txn
            WHERE student_id=? AND item=? COLLATE NOCASE AND returned_qty < qty
            ORDER BY borrow_ts ASC
            ''',
            (student_id, normalize_component_name(item)),
        ).fetchall()
        for row_id, qty, returned_qty in rows:
            open_qty = int(qty) - int(returned_qty)
            if open_qty <= 0:
                continue
            used = min(open_qty, remaining)
            cur.execute("UPDATE borrow_txn SET returned_qty = returned_qty + ? WHERE id=?", (used, row_id))
            applied += used
            remaining -= used
            if remaining <= 0:
                break
        cur.execute("COMMIT")
    except Exception:
        try:
            cur.execute("ROLLBACK")
        except Exception:
            pass
        logging.exception("apply_return_fifo failed")
        return 0, int(return_qty)
    finally:
        conn.close()
    return applied, remaining


def export_components_excel() -> bytes | None:
    if Workbook is None:
        return None
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "components"
    sheet.append(["name", "quantity", "location", "locker"])
    for row in list_all_components():
        sheet.append(list(row))
    bio = BytesIO()
    workbook.save(bio)
    return bio.getvalue()


def export_students_excel() -> bytes | None:
    if Workbook is None:
        return None
    conn = connect_lab()
    rows = conn.execute("SELECT student_id, name, COALESCE(email, '') FROM students ORDER BY name").fetchall()
    conn.close()
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "students"
    sheet.append(["student_id", "name", "email"])
    for row in rows:
        sheet.append(list(row))
    bio = BytesIO()
    workbook.save(bio)
    return bio.getvalue()


def import_components_excel(filepath: Path) -> int:
    if openpyxl is None or not filepath.exists():
        return 0
    workbook = openpyxl.load_workbook(filepath)
    sheet = workbook.active
    headers = [str(cell.value).strip().lower() for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
    index = {name: pos for pos, name in enumerate(headers)}
    required = {"name", "quantity", "location", "locker"}
    if not required.issubset(index):
        return 0
    conn = connect_lab()
    cur = conn.cursor()
    count = 0
    for row in sheet.iter_rows(min_row=2):
        name = str(row[index["name"]].value or "").strip()
        if not name:
            continue
        quantity = int(row[index["quantity"]].value or 0)
        location = str(row[index["location"]].value or "").strip()
        locker = int(row[index["locker"]].value or 0)
        cur.execute(
            '''
            INSERT INTO components(name, quantity, location, locker) VALUES(?,?,?,?)
            ON CONFLICT(name) DO UPDATE SET
                quantity=excluded.quantity,
                location=excluded.location,
                locker=excluded.locker
            ''',
            (normalize_component_name(name), quantity, location, locker),
        )
        count += 1
    conn.commit()
    conn.close()
    return count


def check_and_send_overdue_reminders() -> int:
    settings = get_settings()
    now = _utcnow()
    cutoff = now - dt.timedelta(days=settings.overdue_reminder_min_days)
    conn = connect_lab()
    rows = conn.execute(
        '''
        SELECT
            bt.id,
            bt.borrow_ts,
            bt.due_ts,
            bt.student_name,
            COALESCE(bt.student_email, s.email, ''),
            bt.item,
            bt.qty,
            bt.returned_qty
        FROM borrow_txn bt
        LEFT JOIN students s ON s.student_id = bt.student_id
        WHERE bt.returned_qty < bt.qty
          AND bt.due_ts <= ?
          AND (bt.last_reminder_ts IS NULL OR bt.last_reminder_ts <= ?)
        ''',
        (now.isoformat(timespec="seconds"), cutoff.isoformat(timespec="seconds")),
    ).fetchall()
    conn.close()

    sent = 0
    for row_id, borrow_ts, due_ts, student_name, email, item, qty, returned_qty in rows:
        if not email:
            continue
        outstanding = int(qty) - int(returned_qty)
        subject = f"Lab Reminder: Please return {outstanding}x {item}"
        body = (
            f"Hello {student_name},\n\n"
            f"You still have {outstanding}x {item}.\n"
            f"Borrow date: {borrow_ts}\n"
            f"Due date: {due_ts}\n\n"
            f"Please return the component as soon as possible.\n\n"
            f"Regards,\nELA"
        )
        ok, _ = send_simple_email(email, subject, body)
        if ok:
            sent += 1
            conn2 = connect_lab()
            conn2.execute(
                "UPDATE borrow_txn SET last_reminder_ts=?, reminder_count = reminder_count + 1 WHERE id=?",
                (now.isoformat(timespec="seconds"), row_id),
            )
            conn2.commit()
            conn2.close()
    return sent


def overdue_reminder_loop() -> None:
    import time
    settings = get_settings()
    while True:
        try:
            sent = check_and_send_overdue_reminders()
            if sent:
                logging.info("Overdue reminders sent: %s", sent)
        except Exception:
            logging.exception("overdue_reminder_loop error")
        time.sleep(settings.overdue_check_interval_hours * 3600)
